
from openpyxl import load_workbook
from openpyxl import Workbook

workbook = load_workbook('./res/openpyxl_tst.xlsx');
print(workbook.sheetnames);
sheet = workbook[workbook.sheetnames[0]];
print(sheet.title);

for row in range(1,5):
    for col in range(65,69): # why 65???
        cell_index = chr(col) + str(row);
        print(sheet[cell_index].value, end='\t')
    print();


