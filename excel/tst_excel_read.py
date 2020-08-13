#!/usr/bin/env python3
# -*- coding:utf-8 -*-

import openpyxl;
import pprint;

'''
{'msg_type':{ 'busi_type':{ 'tract':value, 'items':value},
              ...
             }
}
'''

def tst_excel_read(st_name='Sheet1', wb_name='./data.xlsx'):
    #打开工作簿
    wb = openpyxl.load_workbook(wb_name);
    print(wb.sheetnames);
    #选取工作表
    st = wb[st_name];

    #最大行列数
    max_row = st.max_row;
    max_column = st.max_column;
    print("max_row[%d] max_column[%d]" % (max_row, max_column));

    tstData = {};
    for row in range(2, st.max_row + 1):
        msg_type = st['A' + str(row)].value;
        busi_type = st['B' + str(row)].value;
        busi_kind = st['C' + str(row)].value;
        items = st['D' + str(row)].value;

        tstData.setdefault(msg_type,{});
        tstData[msg_type].setdefault(busi_type,{'tract':0, 'items':0});

        tstData[msg_type][busi_type]['tract'] += 1;
        tstData[msg_type][busi_type]['items'] += int(items);

    
    with open( 'data.txt', 'w' ) as f:
        f.write( pprint.pformat(tstData));

    #关闭工作簿
    wb.close();

def main():
    tst_excel_read();

if __name__ == '__main__':
    main();
