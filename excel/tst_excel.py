# -*- coding:utf-8 -*-

import openpyxl;

""" WorkBook Sheet Cell """

def tst_excel_creatwb(st_name='tst_sheet', wb_name='./tst.xlsx'):
    #创建工作簿
    wb = openpyxl.Workbook();
    #创建工作表
    st = wb.create_sheet(st_name);
    #保存工作簿
    wb.save(wb_name);
    wb.close();

def tst_excel_creatst(st_name='tst_sheet', wb_name='./tst.xlsx'):
    wb = openpyxl.load_workbook(wb_name);
    st = wb.create_sheet(st_name);
    wb.save(wb_name);
    wb.close();

def tst_excel_read(st_name='tst_sheet', wb_name='./tst.xlsx'):
    #打开工作簿
    wb = openpyxl.load_workbook(wb_name);
    print(wb.sheetnames);
    #选取工作表
    st = wb[st_name];
    #最大行列数
    max_row = st.max_row;
    max_column = st.max_column;
    print("max_row[%d] max_column[%d]" % (max_row, max_column));
    #选取第一行，第一列数据读取
    ce = st.cell(row=1,column=1);
    #读取行
    rowData = list(st.rows);
    #读取列
    columnData = list(st.columns);
    #关闭工作簿
    wb.close();

    print(ce);
    print(rowData);
    print(columnData);

def tst_excel_write(st_name="write_sheet", wb_name="./tst.xlsx"):
    wb = openpyxl.load_workbook(wb_name);
    st = wb[st_name];
    #写入
    data = 3;
    st.cell(row=1,column=1,value=data);
    st.cell(row=1,column=2).value=data;
    st['C1']=data;
    wb.save(wb_name);
    wb.close();

def tst_excel_removest (st_name="write_sheet", wb_name="./tst.xlsx"):
    wb = openpyxl.load_workbook(wb_name);
    st = wb[st_name];
    wb.remove(st);

def main():
    tst_excel_creatwb();
    tst_excel_read();
    tst_excel_creatst('write_sheet');
    tst_excel_write("write_sheet")

if __name__ == '__main__':
    main();