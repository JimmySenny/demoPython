#!/usr/bin/env python3
# -*- coding:utf-8 -*-

import openpyxl;

def tst_excel_write(st_name='write', wb_name='./data.xlsx'):
    #打开工作簿
    wb = openpyxl.Workbook();
    st = wb.active;

    print( st.title );
    st.title = st_name;
    print( wb.get_sheet_names() );

    st['A1'] = 'hello';
    print(st['A1'].value);

    st1 = wb.create_sheet('sheet1');
    for row in range( 1, 10 ):
        st1.append(range(10));

    st2 = wb.create_sheet('sheet2');
    rows = [
            ['Number', 'Batch1', 'Batch2'],
            [1, 11, 111],
            [2, 12, 121],
            [3, 13, 131],
            [4, 14, 141],
            ];

    for row in rows:
        st2.append(row);

    st3 = wb.create_sheet('sheet3');
    for row in range( 2, 10 ):
        for col in range( 1, 5 ):
            st3.cell(row=row, column=col,
                    value=openpyxl.utils.get_column_letter(col));

    wb.save('example_write.xlsx');




def main():
    tst_excel_write();

if __name__ == '__main__':
    main();
